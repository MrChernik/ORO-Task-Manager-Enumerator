import sys
import os
import time
from datetime import datetime 
import subprocess
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QMessageBox, QProgressBar)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QFont, QColor, QIcon
#Да, это делалось с помощью нейросети, И ЧТО? Главное что работает, а раз ты заинтересовался этим проектом, значит он тебе нужен
#Если это глупо, но это работает - это не глупо!©
# Импорты UI
from splash import AnimatedSplashScreen
from main_window import Ui_MainWindow

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import Select
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


class ParserThread(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    progress_stage_signal = pyqtSignal(str)  # Сигнал для смены этапа прогресса
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, url, login, password):
        super().__init__()
        self.url = url
        self.login = login
        self.password = password
        self.driver = None
        self.is_running = True  # Флаг для контроля выполнения
        self.tasks_data = []  # Храним данные задач
        self.current_stage = "pages"  # Этап: pages или tasks
        
    def stop(self):
        """Остановка парсинга"""
        self.is_running = False
        if self.driver:
            self.driver.quit()
    
    def run(self):
        try:
            self.login_to_task_manager()
        except Exception as e:
            self.log_signal.emit(f"Критическая ошибка: {str(e)}")
            self.finished_signal.emit(False, str(e))
    
    def login_to_task_manager(self):
        try:
            if not self.is_running:
                return
                
            self.log_signal.emit("Инициализация браузера...")
            self.driver = webdriver.Chrome()
            
            self.log_signal.emit("Переход по ссылке...")
            self.driver.get(self.url)
            
            wait = WebDriverWait(self.driver, 10)
            
            self.log_signal.emit("Ввод логина...")
            username_field = wait.until(
                EC.presence_of_element_located((By.ID, "prependedInput"))
            )
            username_field.clear()
            username_field.send_keys(self.login)
            
            self.log_signal.emit("Ввод пароля...")
            password_field = self.driver.find_element(By.ID, "prependedInput2")
            password_field.clear()
            password_field.send_keys(self.password)
            
            self.log_signal.emit("Авторизация...")
            submit_button = self.driver.find_element(By.ID, "_submit")
            submit_button.click()
            
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "grid-body")))
            self.log_signal.emit("Авторизация выполнена успешно!")
            
            # Установка максимального количества записей на странице (100)
            self.set_max_records_per_page(self.driver, wait)
            
            # Этап 1: Сбор данных со всех страниц
            self.current_stage = "pages"
            self.progress_stage_signal.emit("pages")
            self.log_signal.emit("Начинаю сбор основных данных...")
            all_tasks_data = self.collect_all_pages_data(self.driver, wait)
            
            if not self.is_running:
                self.save_partial_data(all_tasks_data, "partial_initial")
                return
                
            # Создание промежуточного Excel файла
            self.create_initial_excel_file(all_tasks_data)
            
            self.log_signal.emit(f"Основные данные сохранены в файл ORO_initial.xlsx на рабочем столе!")
            self.log_signal.emit(f"Всего собрано задач: {len(all_tasks_data)}")
            
            # Этап 2: Сбор комментариев и инициатора
            self.current_stage = "tasks"
            self.progress_stage_signal.emit("tasks")
            self.progress_signal.emit(0)  # Сброс прогресса
            
            self.log_signal.emit("Начинаю сбор комментариев и данных об инициаторе...")
            all_tasks_with_comments = self.collect_comments_and_initiator_for_all_tasks(self.driver, wait, all_tasks_data)
            
            if not self.is_running:
                self.save_partial_data(all_tasks_with_comments, "partial_final")
                return
                
            # Создание финального Excel файла с комментариями и инициатором
            self.create_final_excel_file(all_tasks_with_comments)
            
            self.log_signal.emit(f"Финальные данные с комментариями и инициатором сохранены в файл ORO.xlsx на рабочем столе!")
            self.finished_signal.emit(True, "Парсинг завершен успешно!")
            
        except Exception as e:
            self.log_signal.emit(f"Ошибка при парсинге: {str(e)}")
            self.finished_signal.emit(False, str(e))
        finally:
            if self.driver:
                self.driver.quit()

    def save_partial_data(self, tasks_data, file_prefix):
        """Сохранение частичных данных при прерывании"""
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(desktop_path, f"{file_prefix}_{timestamp}.xlsx")
            
            if tasks_data:
                df = pd.DataFrame(tasks_data)
                df.to_excel(file_path, sheet_name='Задачи', index=False)
                self.log_signal.emit(f"Сохранены частичные данные в файл: {file_path}")
            else:
                self.log_signal.emit("Нет данных для сохранения")
        except Exception as e:
            self.log_signal.emit(f"Ошибка при сохранении частичных данных: {str(e)}")

    def set_max_records_per_page(self, driver, wait):
        """Установка максимального количества записей на странице (100)"""
        try:
            self.log_signal.emit("Установка количества записей на странице (100)...")
            records_button = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn.dropdown-toggle[data-toggle='dropdown']"))
            )
            records_button.click()
            
            time.sleep(1)
            
            option_100 = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a.dropdown-item[data-size='100']"))
            )
            option_100.click()
            
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "grid-body")))
            time.sleep(3)
            
            self.log_signal.emit("Количество записей на странице установлено на 100")
            
        except Exception as e:
            self.log_signal.emit(f"Не удалось установить количество записей на 100: {e}")

    def collect_all_pages_data(self, driver, wait):
        """Сбор данных со всех страниц (без комментариев и инициатора)"""
        all_tasks_data = []
        current_page = 1
        total_pages = self.get_total_pages(driver)
        
        self.log_signal.emit(f"Всего страниц для обработки: {total_pages}")
        
        while self.is_running:
            self.log_signal.emit(f"Обрабатываю страницу {current_page}...")
            
            page_tasks = self.extract_tasks_data(driver)
            all_tasks_data.extend(page_tasks)
            
            # Обновляем прогресс для страниц
            if total_pages > 0:
                progress = int((current_page / total_pages) * 100)
                self.progress_signal.emit(progress)
            
            self.log_signal.emit(f"Собрано задач на странице {current_page}: {len(page_tasks)}")
            
            if not self.has_next_page(driver):
                self.log_signal.emit("Достигнута последняя страница")
                break
            
            if self.go_to_next_page(driver, wait):
                current_page += 1
            else:
                self.log_signal.emit("Не удалось перейти на следующую страницу")
                break
                
            time.sleep(2)
        
        self.progress_signal.emit(100)  # Завершаем этап страниц
        return all_tasks_data

    def get_total_pages(self, driver):
        """Получение общего количества страниц"""
        try:
            # Ищем элемент с информацией о пагинации
            pagination_info = driver.find_elements(By.CSS_SELECTOR, ".grid-pagination-total")
            if pagination_info:
                text = pagination_info[0].text
                # Пытаемся извлечь число из текста типа "1 - 100 из 1234"
                import re
                numbers = re.findall(r'\d+', text)
                if len(numbers) >= 3:
                    total_items = int(numbers[2])
                    items_per_page = 100  # Мы установили 100 на странице
                    total_pages = (total_items + items_per_page - 1) // items_per_page
                    return total_pages
            return 10  # Возвращаем приблизительное значение по умолчанию
        except:
            return 10  # Возвращаем приблизительное значение по умолчанию

    def has_next_page(self, driver):
        """Проверка наличия следующей страницы"""
        try:
            next_button = driver.find_element(By.CSS_SELECTOR, "button[data-grid-pagination-trigger][data-grid-pagination-direction='next']")
            return "disabled" not in next_button.get_attribute("class")
        except:
            return False

    def go_to_next_page(self, driver, wait):
        """Переход на следующую страницу"""
        try:
            next_button = driver.find_element(By.CSS_SELECTOR, "button[data-grid-pagination-trigger][data-grid-pagination-direction='next']")
            
            if "disabled" not in next_button.get_attribute("class"):
                next_button.click()
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "grid-body")))
                time.sleep(2)
                return True
            return False
        except Exception as e:
            self.log_signal.emit(f"Ошибка при переходе на следующую страницу: {e}")
            return False

    def extract_tasks_data(self, driver):
        """Извлечение данных из таблицы задач на текущей странице"""
        tasks_data = []
        
        task_rows = driver.find_elements(By.CSS_SELECTOR, "tr.grid-row")
        
        self.log_signal.emit(f"Найдено строк на странице: {len(task_rows)}")
        
        for i, row in enumerate(task_rows):
            if not self.is_running:
                break
                
            try:
                task_data = self.extract_task_from_row(row)
                if task_data:
                    tasks_data.append(task_data)
                    
            except Exception as e:
                self.log_signal.emit(f"Ошибка при обработке строки {i}: {e}")
                continue
        
        return tasks_data

    def extract_task_from_row(self, row):
        """Извлечение данных из одной строки задачи"""
        try:
            subject_element = row.find_element(By.CSS_SELECTOR, "td.grid-body-cell-subject")
            
            task_name = ""
            try:
                task_name_element = subject_element.find_element(By.CSS_SELECTOR, "span.task-state-default")
                task_name = task_name_element.text.strip()
            except:
                try:
                    task_name_element = subject_element.find_element(By.CSS_SELECTOR, "a.task-subject")
                    task_name = task_name_element.text.strip()
                except:
                    task_name = subject_element.text.strip()
            
            view_link_element = row.find_element(By.CSS_SELECTOR, "a[title='Просмотр']")
            task_link = view_link_element.get_attribute("href")

            try:
                created_element = row.find_element(By.CSS_SELECTOR, "td.grid-body-cell-createdAt")
                created_date = created_element.text.strip()
            except Exception: created_date = 'N/A'
            
            deadline_element = row.find_element(By.CSS_SELECTOR, "td.grid-body-cell-deadline")
            deadline = deadline_element.text.strip()
            
            owner_element = row.find_element(By.CSS_SELECTOR, "td.grid-body-cell-ownerName")
            owner = owner_element.text.strip()
            
            status_element = row.find_element(By.CSS_SELECTOR, "td.grid-body-cell-statusLabel")
            status = status_element.text.strip()
            
            task_data = {
                "Название задачи": task_name,
                "Ссылка на задачу": task_link,
                "Дата постановки задачи": created_date,
                "Дедлайн задачи": deadline,
                "Инициатор": "",
                "Ответственный по задаче": owner,
                "Статус задачи": status,
                "Автор последнего комментария": "",
                "Последний комментарий": "",
                "Кто тормозит": "",
                "Причина остановки/задержки в выполнении задачи": ""
            }
            
            return task_data
            
        except Exception as e:
            self.log_signal.emit(f"Ошибка при извлечении данных из строки: {e}")
            return None

    def collect_comments_and_initiator_for_all_tasks(self, driver, wait, tasks_data):
        """Сбор комментариев и данных об инициатора для всех задач"""
        tasks_with_comments = []
        total_tasks = len(tasks_data)
        
        for i, task in enumerate(tasks_data, 1):
            if not self.is_running:
                break
                
            self.log_signal.emit(f"Обрабатываю задачу {i}/{total_tasks}: {task['Название задачи'][:50]}...")
            
            try:
                author, comment, initiator = self.get_last_comment_and_initiator(driver, task['Ссылка на задачу'], wait)
                task['Автор последнего комментария'] = author
                task['Последний комментарий'] = comment
                task['Инициатор'] = initiator
                tasks_with_comments.append(task)
                
                # Обновляем прогресс для задач
                progress = int((i / total_tasks) * 100)
                self.progress_signal.emit(progress)
                
                time.sleep(1)
                
            except Exception as e:
                self.log_signal.emit(f"Ошибка при обработке задачи {i}: {e}")
                task['Автор последнего комментария'] = f"Ошибка: {str(e)}"
                task['Последний комментарий'] = ""
                task['Инициатор'] = f"Ошибка: {str(e)}"
                tasks_with_comments.append(task)
                continue
        
        return tasks_with_comments

    def get_last_comment_and_initiator(self, driver, task_url, wait):
        """Получение последнего комментария и инициатора из задачи по URL"""
        try:
            original_window = driver.current_window_handle
            
            driver.execute_script("window.open('');")
            driver.switch_to.window(driver.window_handles[-1])
            driver.get(task_url)
            
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "responsive-block")))
            time.sleep(2)
            
            initiator = self.get_initiator(driver)
            author, comment = self.get_last_comment_from_page(driver, wait)
            
            self.log_signal.emit(f"  Инициатор: {initiator}")
            self.log_signal.emit(f"  Комментарий от {author}: {comment[:50]}...")
            
            driver.close()
            driver.switch_to.window(original_window)
            
            return author, comment, initiator
            
        except Exception as e:
            self.log_signal.emit(f"  Ошибка при получении данных: {e}")
            try:
                driver.switch_to.window(original_window)
            except:
                pass
            return "Ошибка", f"Не удалось получить комментарий: {str(e)}", f"Ошибка: {str(e)}"

    def get_initiator(self, driver):
        """Получение данных об инициаторе (постановщике) задачи"""
        try:
            # Способ 1
            try:
                initiator_label = driver.find_element(By.XPATH, "//label[contains(@class, 'attribute-item__term') and contains(text(), 'Постановщик')]")
                initiator_item = initiator_label.find_element(By.XPATH, "./..")
                initiator_value = initiator_item.find_element(By.CSS_SELECTOR, ".select2-result-label-title")
                return initiator_value.text.strip()
            except:
                pass

            # Способ 2
            try:
                attribute_items = driver.find_elements(By.CSS_SELECTOR, ".attribute-item")
                for item in attribute_items:
                    try:
                        label = item.find_element(By.CSS_SELECTOR, ".attribute-item__term")
                        if "Постановщик" in label.text:
                            value = item.find_element(By.CSS_SELECTOR, ".select2-result-label-title")
                            return value.text.strip()
                    except:
                        continue
            except:
                pass

            # Способ 3
            try:
                elements_with_text = driver.find_elements(By.XPATH, "//*[contains(text(), 'Постановщик')]")
                for element in elements_with_text:
                    try:
                        parent = element.find_element(By.XPATH, "./ancestor::div[contains(@class, 'attribute-item')][1]")
                        value = parent.find_element(By.CSS_SELECTOR, ".select2-result-label-title")
                        return value.text.strip()
                    except:
                        continue
            except:
                pass

            # Способ 4
            try:
                participants_blocks = driver.find_elements(By.CSS_SELECTOR, ".responsive-block")
                for block in participants_blocks:
                    try:
                        if "Постановщик" in block.text:
                            initiator_elements = block.find_elements(By.CSS_SELECTOR, ".select2-result-label-title")
                            if len(initiator_elements) >= 2:
                                return initiator_elements[1].text.strip()
                    except:
                        continue
            except:
                pass

            # Способ 5
            try:
                all_titles = driver.find_elements(By.CSS_SELECTOR, ".select2-result-label-title")
                for i, title in enumerate(all_titles):
                    parent_container = title.find_element(By.XPATH, "./ancestor::div[contains(@class, 'attribute-item')][1]")
                    if "Постановщик" in parent_container.text:
                        return title.text.strip()
            except:
                pass

            return "Неизвестный инициатор"
            
        except Exception as e:
            return "Ошибка при поиске инициатора"
    
    def get_last_comment_from_page(self, driver, wait):
        """Получение последнего комментария со страницы задачи"""
        try:
            comment_items = driver.find_elements(By.CSS_SELECTOR, "li.comment-item")
            
            if comment_items:
                last_comment = comment_items[-1]
                
                try:
                    author_element = last_comment.find_element(By.CSS_SELECTOR, "a.user")
                    author = author_element.text.strip()
                except:
                    author = "Неизвестный автор"
                
                try:
                    comment_body = last_comment.find_element(By.CSS_SELECTOR, "div.comment-body")
                    comment_text = comment_body.text.strip()
                except:
                    comment_text = "Текст комментария не найден"
                
                return author, comment_text
                
            else:
                return "Нет комментариев", ""
            
        except Exception as e:
            return "Ошибка", f"Не удалось получить комментарий: {str(e)}"

    def create_initial_excel_file(self, tasks_data):
        """Создание промежуточного Excel файла без комментариев и инициатора"""
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, "ORO_initial.xlsx")
        
        if os.path.exists(file_path):
            os.remove(file_path)
        
        df = pd.DataFrame(tasks_data)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Задачи', index=False)
            
            worksheet = writer.sheets['Задачи']
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 60
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 25
            worksheet.column_dimensions['F'].width = 30
            worksheet.column_dimensions['G'].width = 20
            worksheet.column_dimensions['H'].width = 25
            worksheet.column_dimensions['I'].width = 40
            worksheet.column_dimensions['J'].width = 25
            worksheet.column_dimensions['K'].width = 40

    def create_final_excel_file(self, tasks_data):
        """Создание финального Excel файла с комментариями, инициатором и цветовым форматированием"""
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
         
        now = datetime.now()  
        dateprefix = now.strftime('%d.%m.%Y') 
        file_path = os.path.join(desktop_path, dateprefix+"_ORO_.xlsx")
        
        if os.path.exists(file_path):
            os.remove(file_path)
        
        df = pd.DataFrame(tasks_data)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Задачи', index=False)
            
            worksheet = writer.sheets['Задачи']
            
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 60
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 25
            worksheet.column_dimensions['F'].width = 30
            worksheet.column_dimensions['G'].width = 20
            worksheet.column_dimensions['H'].width = 25
            worksheet.column_dimensions['I'].width = 40
            worksheet.column_dimensions['J'].width = 25
            worksheet.column_dimensions['K'].width = 40
            
            for row_idx, task in enumerate(tasks_data, start=2):
                link_cell = worksheet.cell(row=row_idx, column=2)
                link_url = task['Ссылка на задачу']
                link_cell.hyperlink = link_url
                link_cell.value = link_url
                link_cell.style = 'Hyperlink'
            
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            
            for row_idx, task in enumerate(tasks_data, start=2):
                status = task.get("Статус задачи", "").lower()
                
                if any(word in status for word in ['условн']):
                    fill = blue_fill
                elif any(word in status for word in ['завершена']):
                    fill = green_fill 
                else:
                    fill = yellow_fill
                
                for col_idx in range(1, 12):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fill


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.parser_thread = None
        self.current_stage = None  # Текущий этап: pages или tasks
        
        # Устанавливаем иконку приложения
        self.setWindowIcon(QIcon('favicon.ico'))
        
        self.apply_custom_styles()
        self.setup_connections()
        self.set_default_values()
        self.setup_title_animation()
        
    def apply_custom_styles(self):
        """Применение кастомных стилей с улучшенным дизайном кнопки"""
        # Обновляем стили с улучшенной кнопкой
        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #404041;
                color: #b48c50;
            }
            QLineEdit, QTextEdit {
                background-color: #505052;
                color: #b48c50;
                border: 1px solid #b48c50;
                border-radius: 3px;
                padding: 5px;
                font-family: Times New Roman;
                font-size: 14px;
            }
            QPushButton {
                background-color: #505052;
                color: #b48c50;
                border: 2px solid #b48c50;
                border-radius: 10px;
                padding: 8px 15px;
                font-family: Times New Roman;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #b48c50;
                color: #404041;
                border: 2px solid #b48c50;
            }
            QPushButton:pressed {
                background-color: #a07c40;
                color: #404041;
                border: 2px solid #a07c40;
            }
            QPushButton:disabled {
                background-color: #606062;
                color: #807c70;
                border: 2px solid #807c70;
            }
            QLabel {
                font-family: Times New Roman;
                font-size: 14px;
                color: #b48c50;
            }
            QProgressBar {
                border: 1px solid #b48c50;
                border-radius: 3px;
                text-align: center;
                background-color: #505052;
                color: #b48c50;
                font-family: Times New Roman;
                font-size: 12px;
            }
            QProgressBar::chunk {
                background-color: #b48c50;
                border-radius: 2px;
            }
        """)
        
        # Простой стиль для заголовка без градиента
        title_style = """
            QLabel {
                color: #b48c50;
                font-family: Times New Roman;
                font-size: 40pt;
                font-weight: bold;
            }
        """
        self.label.setStyleSheet(title_style)
        
    def setup_connections(self):
        """Подключение сигналов к слотам"""
        self.parseButton.clicked.connect(self.start_parsing)
        
    def set_default_values(self):
        """Установка значений по умолчанию"""
        default_url = "Вставь сюда свою ссылку/креды по умолчанию для удобства, и раскомментируй три строки ниже"
        #self.lineEdit.setText(default_url)
        #self.loginInput.setText("")
        #self.passwordInput.setText("")
        
    def setup_title_animation(self):
        """Настройка анимации плавного исчезания и появления заголовка"""
        self.animation_sequence = QPropertyAnimation(self.label, b"windowOpacity")
        self.animation_sequence.setDuration(6000)  # Полный цикл: появление + исчезание
        self.animation_sequence.setKeyValueAt(0, 1.0)
        self.animation_sequence.setKeyValueAt(0.5, 0.1)  # Минимальная прозрачность в середине
        self.animation_sequence.setKeyValueAt(1, 1.0)    # Возврат к полной прозрачности
        self.animation_sequence.setLoopCount(-1)  # Бесконечное повторение
        
        # Запускаем анимацию
        self.animation_sequence.start()
        
    def mouseDoubleClickEvent(self, event):
        """Обработчик двойного клика по форме"""
        QMessageBox.information(self, "Сообщение", "НЕ ВСЁ, ЧТО НЕ РАБОТАЕТ, ТО РЭБ")
        
    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        if self.parser_thread and self.parser_thread.isRunning():
            reply = QMessageBox.question(self, 'Подтверждение закрытия',
                                       'Парсинг еще выполняется. Вы хотите остановить его и сохранить текущие данные?',
                                       QMessageBox.Yes | QMessageBox.No,
                                       QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                self.parser_thread.stop()
                self.parser_thread.wait(5000)  # Ждем до 5 секунд для корректного завершения
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()
        
    def check_dependencies(self):
        """Проверка и установка зависимостей"""
        self.textEdit.setVisible(True)
        self.textEdit.append("Проверка зависимостей...")
        
        if not SELENIUM_AVAILABLE:
            self.textEdit.append("Установка необходимых пакетов...")
            try:
                packages = ["selenium", "pandas", "openpyxl", "PyQt5"]
                for package in packages:
                    self.textEdit.append(f"Установка {package}...")
                    subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                
                self.textEdit.append("Зависимости успешно установлены! Перезапустите программу.")
                QMessageBox.information(self, "Успех", "Зависимости установлены! Перезапустите программу.")
            except Exception as e:
                self.textEdit.append(f"Ошибка установки зависимостей: {str(e)}")
                QMessageBox.critical(self, "Ошибка", f"Не удалось установить зависимости: {str(e)}")
        else:
            self.textEdit.append("Все зависимости установлены!")
    
    def start_parsing(self):
        """Запуск парсинга"""
        url = self.lineEdit.text().strip()
        login = self.loginInput.text().strip()
        password = self.passwordInput.text().strip()
        
        if not url or not login or not password:
            QMessageBox.warning(self, "Ошибка", "Заполните все поля!")
            return
        
        if not SELENIUM_AVAILABLE:
            QMessageBox.warning(self, "Ошибка", "Сначала установите зависимости!")
            return
        
        self.textEdit.setVisible(True)
        self.textEdit.clear()
        self.progressBar.setVisible(True)
        self.progressBar.setValue(0)
        self.parseButton.setEnabled(False)
        
        self.parser_thread = ParserThread(url, login, password)
        self.parser_thread.log_signal.connect(self.update_log)
        self.parser_thread.progress_signal.connect(self.update_progress)
        self.parser_thread.progress_stage_signal.connect(self.update_progress_stage)
        self.parser_thread.finished_signal.connect(self.parsing_finished)
        self.parser_thread.start()
    
    def update_log(self, message):
        """Обновление логов"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.textEdit.append(f"[{timestamp}] {message}")
        # Автопрокрутка к последнему сообщению
        self.textEdit.verticalScrollBar().setValue(
            self.textEdit.verticalScrollBar().maximum()
        )
    
    def update_progress(self, value):
        """Обновление прогресса"""
        self.progressBar.setValue(value)
    
    def update_progress_stage(self, stage):
        """Обновление этапа прогресса"""
        self.current_stage = stage
        if stage == "pages":
            self.progressBar.setFormat("Обработка страниц: %p%")
        elif stage == "tasks":
            self.progressBar.setFormat("Обработка задач: %p%")
            self.progressBar.setValue(0)  # Сброс для нового этапа
    
    def parsing_finished(self, success, message):
        """Завершение парсинга"""
        self.parseButton.setEnabled(True)
        self.progressBar.setVisible(False)
        
        if success:
            QMessageBox.information(self, "Успех", message)
        else:
            QMessageBox.critical(self, "Ошибка", message)


class OROEnumeratorApp:
    def __init__(self):
        self.app = QApplication(sys.argv)
        
        # Устанавливаем иконку приложения
        self.app.setWindowIcon(QIcon('favicon.ico'))
        
        # Установка шрифта по умолчанию
        default_font = QFont("Times New Roman", 14)
        self.app.setFont(default_font)
        
        self.splash = AnimatedSplashScreen()
        self.main_window = MainWindow()
        
    def run(self):
        self.splash.show()
        
        # Переход на главное окно через 4 секунды
        QTimer.singleShot(4000, self.show_main_window)
        
        return self.app.exec_()
    
    def show_main_window(self):
        self.splash.close()
        self.main_window.show()


if __name__ == "__main__":
    enumerator_app = OROEnumeratorApp()
    sys.exit(enumerator_app.run())
